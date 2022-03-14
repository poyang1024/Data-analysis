#!/usr/bin/env python
# coding: utf-8

# In[1]:


print('使用前請先詳閱Readme文件中的使用步驟......')
print()
print('請耐心等候，程式運行完後會自動關閉，於output資料夾匯出所需檔案......')
print()
print('有出現任何錯誤麻煩截圖告知，感謝您的使用......')
print()
print('Excel檔案中的行列標頭及排序請勿做更動，如有需進行更動，請聯絡工程師進行程式系統改版......')
print()
print('程式運作中，請勿開啟相關Excel檔案，否則會導致權限問題使程式無法正常運行......')


# In[2]:


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
def reset_col(filename):
	wb=load_workbook(filename)
	for sheet in wb.sheetnames:
		ws=wb[sheet]
		df=pd.read_excel(filename,sheet).fillna('-')
		df.loc[len(df)]=list(df.columns)						#把標题行附加到最后一行
		for col in df.columns:				
			index=list(df.columns).index(col)					#列序號
			letter=get_column_letter(index+1)					#列字母
			collen=df[col].apply(lambda x:len(str(x).encode())).max()	#獲取這一列長度的最大值
			ws.column_dimensions[letter].width=collen*0.3+10

	wb.save(filename)


# In[3]:


def percentage(part, whole):
  percentage = 100 * float(part)/float(whole)
  percentageintwosite = round(percentage,2)
  return str(percentageintwosite) + "%"


# In[4]:


from tqdm import tqdm
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBooking = pd.read_excel("input/input.xlsx",sheet_name="訂單明細").fillna('')
df = KindFoodBooking
KindFoodBookingNOKOL = df.loc[KindFoodBooking['會員標籤'].str.contains('KOL') == False]
dfN = KindFoodBookingNOKOL

dftotal = dfN['商品售價']*dfN['數量']
dftotal.name = '加總 - 商品售價'

dfN = pd.concat([dfN,dftotal],axis='columns')

KindFoodBookingHub = dfN.groupby(['SKU']).agg({'加總 - 商品售價':['sum']})

KindFoodBookingHub.name = 'Hub'                            
KindFoodBookingsum = KindFoodBookingHub.sum()

KindFoodBookingsum.name = '總和'

KindFoodBookingHub = KindFoodBookingHub.append(KindFoodBookingsum)


with pd.ExcelWriter("output/KindFoodBookingAnalysis.xlsx") as writer:
  KindFoodBookingNOKOL.to_excel(writer,sheet_name="訂單明細",index=None)
  KindFoodBookingHub.to_excel(writer,sheet_name="數據分析")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[5]:


KindFoodBookingcoupon = pd.read_excel("input/input.xlsx",sheet_name="訂單明細")
dfcoupon = KindFoodBookingcoupon
KindFoodBookingNOKOLcoupon = dfcoupon.loc[KindFoodBooking['會員標籤'].str.contains('KOL') == False]
dfNcoupon = KindFoodBookingNOKOLcoupon

dfNcoupon = dfNcoupon.drop_duplicates('訂單編號','first')


KindFoodBookingHubcoupon = dfNcoupon['優惠券名稱'].value_counts()

KindFoodBookingHubcoupon = KindFoodBookingHubcoupon.rename_axis('列標籤').reset_index(name='計數 - 優惠券名稱')

KindFoodBookingHubcoupon.reset_index(inplace=True, drop=True)

KindFoodBookingsumcoupon = KindFoodBookingHubcoupon['計數 - 優惠券名稱'].sum()

dfcoupon = pd.DataFrame([['總和',KindFoodBookingsumcoupon]],columns=['列標籤','計數 - 優惠券名稱'])

KindFoodBookingHubcoupon = KindFoodBookingHubcoupon.append(dfcoupon,ignore_index=True)

KindFoodBookingHubcoupon = KindFoodBookingHubcoupon.sort_values(by=['計數 - 優惠券名稱'],ascending = False)


# In[6]:


dfbonus = KindFoodBookingNOKOLcoupon
dfbonus = dfbonus.drop_duplicates('訂單編號','first')

KindFoodBookingEarnbonus = dfbonus['可獲得紅利'].sum()
KindFoodBookingUsebonus = dfbonus['紅利折抵'].sum()
KindFoodBookingnumcount = dfbonus['訂單編號'].value_counts()


dfbonus['紅利折抵'] = dfbonus['紅利折抵'].astype(str)

KindFoodBookingnumcountwithoutzero = dfbonus.loc[dfbonus['紅利折抵'].str.find('0') == False]
dfnumcountwithoutzero = KindFoodBookingnumcountwithoutzero['訂單編號'].value_counts()

dfnumcountwithoutzeroTRUE = len(KindFoodBookingnumcount.index) - len(dfnumcountwithoutzero.index)

dfdividend = pd.DataFrame([['總發放紅利',KindFoodBookingEarnbonus],
              ['紅利折抵',KindFoodBookingUsebonus],
              ['紅利使用百分比',percentage(KindFoodBookingUsebonus,KindFoodBookingEarnbonus)],
              ['總訂單數量',len(KindFoodBookingnumcount.index)],
              ['使用紅利訂單數量',dfnumcountwithoutzeroTRUE],
              ['回頭客比例',percentage(dfnumcountwithoutzeroTRUE,len(KindFoodBookingnumcount.index))]],columns=['Title','數值'])


# In[7]:


dftrans = KindFoodBookingNOKOL
dftrans = dftrans.drop_duplicates('訂單編號','first')

KindFoodBookingtransHub = dftrans.groupby(['出貨方式'])
KindFoodBookingtransHub = KindFoodBookingtransHub.size().reset_index(name='計數 - 出貨方式')

KindFoodBookingtransHub.name = 'Hub'                            
KindFoodBookingtranssum = KindFoodBookingtransHub.sum()

KindFoodBookingtranssum.name = '總和'

KindFoodBookingtransHub = KindFoodBookingtransHub.append(KindFoodBookingtranssum)

KindFoodBookingtransHub = KindFoodBookingtransHub.sort_values(by=['計數 - 出貨方式'],ascending = False)
KindFoodBookingtransHub.reset_index(inplace=True, drop=True)

KindFoodBookingtransHub.loc[0,'出貨方式'] = '總和'
KindFoodBookingtransHub["物流方式佔比"] = ''


# In[8]:


dfknow = KindFoodBookingNOKOL
dfknow = dfknow.drop_duplicates('訂單編號','first')

KindFoodBookingknowHub = dfknow.groupby(['額外資訊'])

KindFoodBookingknowHub = KindFoodBookingknowHub.size().reset_index(name='計數')

KindFoodBookingknowHub = KindFoodBookingknowHub.sort_values(by=['計數'],ascending = False)

KindFoodBookingknowHub.reset_index(inplace=True, drop=True)

KindFoodBookingknowledgeHub = KindFoodBookingknowHub['額外資訊'].str.replace('如何知道康福先生的呢？: ', '')
KindFoodBookingknowHub = KindFoodBookingknowHub.rename(columns = {'額外資訊':'刪除'})
KindFoodBookingknowHub = pd.concat([KindFoodBookingknowledgeHub , KindFoodBookingknowHub] , axis = 1)
KindFoodBookingknowHub = KindFoodBookingknowHub.drop(columns={'刪除'})
KindFoodBookingknowHub = KindFoodBookingknowHub.drop([0])
KindFoodBookingknowHub = KindFoodBookingknowHub.rename(columns = {'額外資訊':'如何知道康福的呢?'})


# In[9]:


KindFoodBookingcount = pd.read_excel("output/KindFoodBookingAnalysis.xlsx",sheet_name="數據分析")
dfcount = KindFoodBookingcount

dfcount = dfcount.drop([0,1])

dfcount = dfcount.rename(columns={"Unnamed: 0":"列標籤"})
dfcount["商品營業額佔比"] = ''
dfcount = dfcount.sort_values(by=['加總 - 商品售價'],ascending = False)
dfcount.reset_index(inplace=True, drop=True)

dfstoreindex = dfcount

dfnew = dfcount["列標籤"].str.slice(0,2)
dfnew.name = '標籤'


dfcount = dfcount.merge(dfnew ,how='inner', left_index=True, right_index=True)
dfcount = dfcount.drop(columns={'列標籤'})
dfcount = dfcount.rename(columns = {'標籤':'列標籤'})
dfcount = dfcount.reindex(columns = dfstoreindex.columns)

KindFoodBookingcountHub = dfcount.groupby(['列標籤']).agg({'加總 - 商品售價':['sum']})

KindFoodBookingcountHub.name = 'Hub'       

with pd.ExcelWriter("output/KindFoodBookingAnalysis.xlsx") as writer:
  KindFoodBookingNOKOL.to_excel(writer,sheet_name="訂單明細",index=None)
  KindFoodBookingcountHub.to_excel(writer,sheet_name="分類數據分析")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[10]:


KindFoodBookingcountall = pd.read_excel("output/KindFoodBookingAnalysis.xlsx",sheet_name="分類數據分析")
dfcountall = KindFoodBookingcountall

dfcountall= dfcountall.drop([0,1])

dfcountall = dfcountall.rename(columns={"Unnamed: 0":"列標籤"})
dfcountall["商品營業額佔比"] = ''
dfcountall = dfcountall.sort_values(by=['加總 - 商品售價'],ascending = False)
dfcountall.reset_index(inplace=True, drop=True)

with pd.ExcelWriter("output/KindFoodBookingAnalysis.xlsx") as writer:
  KindFoodBookingNOKOL.to_excel(writer,sheet_name="訂單明細",index=None)
  dfstoreindex.to_excel(writer,sheet_name="數據分析",index=None)
  dfcountall.to_excel(writer,sheet_name="分類數據分析",index=None)
  KindFoodBookingHubcoupon.to_excel(writer,sheet_name="優惠卷數據分析",index=None)
  dfdividend.to_excel(writer,sheet_name="紅利與回頭客數據分析",index=None)
  KindFoodBookingtransHub.to_excel(writer,sheet_name="物流方式占比數據分析",index=None)
  KindFoodBookingknowHub.to_excel(writer,sheet_name="如何知道康福數據分析",index=None)
tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[11]:


import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import colors, Font, Fill, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment

workbook = load_workbook("output/KindFoodBookingAnalysis.xlsx")
sheet = workbook['數據分析']


sheet.insert_rows(1)
sheet["A1"] = "各商品營業額佔比"
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")

for i in range(3,sheet.max_row+1):
  sheet["C"+ format(i)] = "=TEXT(B"+ format(i) + "/B"+ format(3) + ",\"0.00%\")"
  sheet["C"+ format(i)].alignment = Alignment(horizontal="right", vertical="center")

sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)


workbook.save(filename = "output/KindFoodBookingAnalysis.xlsx")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[12]:


sheetcount = workbook['分類數據分析']

sheetcount.insert_rows(1)
sheetcount["A1"] = "各商品營業額佔比"
sheetcount["A1"].alignment = Alignment(horizontal="center", vertical="center")

for i in range(3,sheetcount.max_row+1):
  sheetcount["C"+ format(i)] = "=TEXT(B"+ format(i) + "/B"+ format(3) + ",\"0.00%\")"
  sheetcount["C"+ format(i)].alignment = Alignment(horizontal="right", vertical="center")

sheetcount.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

workbook.save(filename = "output/KindFoodBookingAnalysis.xlsx")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[13]:


sheetbonus = workbook['紅利與回頭客數據分析']
sheetbonus.insert_rows(1)
sheetbonus["A1"] = "紅利/回頭客分析"
sheetbonus["A1"].alignment = Alignment(horizontal="center", vertical="center")

for i in range(3,sheetcount.max_row+1):
  sheetbonus["B"+ format(i)].alignment = Alignment(horizontal="right", vertical="center")

sheetbonus.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

workbook.save(filename = "output/KindFoodBookingAnalysis.xlsx")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[14]:


sheettrans = workbook['物流方式占比數據分析']
sheettrans.insert_rows(1)
sheettrans["A1"] = "物流方式佔比"
sheettrans["A1"].alignment = Alignment(horizontal="center", vertical="center")

for i in range(3,sheettrans.max_row+1):
  sheettrans["C"+ format(i)] = "=TEXT(B"+ format(i) + "/B"+ format(3) + ",\"0.00%\")"
  sheettrans["C"+ format(i)].alignment = Alignment(horizontal="right", vertical="center")

sheettrans.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)

workbook.save(filename = "output/KindFoodBookingAnalysis.xlsx")

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingAnalysis.xlsx")


# In[ ]:




