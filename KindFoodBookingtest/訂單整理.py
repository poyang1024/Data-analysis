#!/usr/bin/env python
# coding: utf-8

# In[1]:


print('使用前請先詳閱Readme文件中的使用步驟......')
print()
print('請耐心等候，程式運行完後會自動關閉，於output資料夾匯出所需檔案......')
print()
print('有出現任何錯誤麻煩截圖告知，感謝您的使用......')
print()
print('input資料夾中的雲端發票開立空白頁請勿移除與使用，程式運行完之匯出檔案都於output資料夾內，請使用output資料夾內中的匯出檔案，謝謝......')
print()
print('Excel檔案中的行列標頭及排序請勿做更動，如有需進行更動，請聯絡工程師進行程式系統改版......')
print()
print('程式運作中，請勿開啟相關Excel檔案，否則會導致權限問題使程式無法正常運行......')


# In[2]:


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
def reset_col(filename):
	wb=load_workbook(filename)
	for sheet in wb.sheetnames:
		ws=wb[sheet]
		df=pd.read_excel(filename,sheet).fillna('-')
		df.loc[len(df)]=list(df.columns)						#把標题行附加到最后一行
		for col in df.columns:				
			index=list(df.columns).index(col)					#列序號
			letter=get_column_letter(index+1)					#列字母
			collen=df[col].apply(lambda x:len(str(x).encode())).max()	#獲取這一列長度的最大值
			ws.column_dimensions[letter].width=collen*0.3+10

	wb.save(filename)


# In[3]:


from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBooking = pd.read_excel("input/input.xlsx",sheet_name="訂單明細").fillna('')

df = KindFoodBooking

KindFoodBookingVIP = df.loc[KindFoodBooking['會員標籤'].str.contains('staff|KOL|經銷')]
KindFoodBookingOrdinary = df.loc[KindFoodBooking['會員標籤'].str.contains('staff|KOL|經銷') == False]

KindFoodBookingVIP.loc[:,'訂單編號'] = '#' + KindFoodBookingVIP.loc[:,'訂單編號']



KindFoodBookingTidy = pd.concat([ KindFoodBookingOrdinary,KindFoodBookingVIP],axis='index')
KindFoodBookingTidy = KindFoodBookingTidy.reindex(KindFoodBooking.index)
KindFoodBookingTidy['收件人電話'] = KindFoodBookingTidy['收件人電話'].apply(str)
KindFoodBookingTidy.loc[:,'收件人電話'] = '0' + KindFoodBookingTidy.loc[:,'收件人電話']


KindFoodBookingHub = KindFoodBooking.groupby(['SKU']).agg({'數量':['sum']})
KindFoodBookingHub.name = 'Hub'                            
KindFoodBookingsum = KindFoodBookingHub.sum()

KindFoodBookingsum.name = '總和'

KindFoodBookingHub = KindFoodBookingHub.append(KindFoodBookingsum)

with pd.ExcelWriter("output/KindFoodBookingTidy.xlsx") as writer:
  KindFoodBookingTidy.to_excel(writer,sheet_name="訂單明細",index=None)
  KindFoodBookingHub.to_excel(writer,sheet_name="樞紐分析")
tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingTidy.xlsx")    


# In[4]:


from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBooking = pd.read_excel("output/KindFoodBookingTidy.xlsx",sheet_name="訂單明細")
df = KindFoodBooking
KindFoodBookingFamily = df.loc[KindFoodBooking['出貨方式'].str.contains('全家')]
KindFoodBookingFamily['收件人電話'] = KindFoodBookingFamily['收件人電話'].apply(str)
KindFoodBookingFamily.loc[:,'收件人電話'] = '0' + KindFoodBookingFamily.loc[:,'收件人電話']

KindFoodBookingFamily.to_excel("output/KindFoodBookingFamily.xlsx",index=None)
tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingFamily.xlsx")


# In[5]:


from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
KindFoodBooking = pd.read_excel("output/KindFoodBookingTidy.xlsx",sheet_name="訂單明細")

df = KindFoodBooking

KindFoodBookingTCat = df.loc[KindFoodBooking['出貨方式'].str.contains('黑貓')]

KindFoodBookingTCat['收件人電話'] = KindFoodBookingTCat['收件人電話'].apply(str)
KindFoodBookingTCat.loc[:,'收件人電話'] = '0' + KindFoodBookingTCat.loc[:,'收件人電話']

KindFoodBookingTCat.to_excel("output/KindFoodBookingTCat.xlsx",index=None)
tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingTCat.xlsx")


# In[6]:


from tqdm import tqdm
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBookingTidy = pd.read_excel("output/KindFoodBookingTidy.xlsx")
KindFoodBookingBill = pd.read_excel("input/雲端發票開立.xlsx",sheet_name="sheet1")


dfB = KindFoodBookingBill
dfT = KindFoodBookingTidy

dfnumTidy = dfT.loc[KindFoodBookingTidy['訂單編號'].str.contains('##') == False]
dfnumTidy = dfnumTidy.drop_duplicates('訂單編號','first')

dfnumTidy = dfnumTidy.reset_index(drop=True)


dfnumTidy = dfnumTidy.rename(columns={'會員名稱':'買受人名稱'})
dfnumTidy = dfnumTidy.rename(columns={'Email':'消費者編號'})
dfnumTidy = dfnumTidy.rename(columns={'收件人電話':'客戶手機'})
dfnumTidy = dfnumTidy.rename(columns={'總額':'小計'})

dfB = dfB.rename(columns = {'買受人名稱':'刪除'})
dfB = dfB.rename(columns = {'消費者編號':'刪除'})
dfB = dfB.rename(columns = {'客戶手機':'刪除'})
dfB = dfB.rename(columns = {'小計':'刪除'})

dfB = dfB.drop(columns={'刪除'})

dfB = dfB.merge(dfnumTidy[['訂單編號','買受人名稱','消費者編號','客戶手機','小計']],on = '訂單編號',how = 'outer')
dfnumTidy = dfnumTidy.rename(columns={'消費者編號':'客戶信箱'})
dfB = dfB.rename(columns = {'客戶信箱':'刪除'})
dfB = dfB.drop(columns={'刪除'})
dfB = dfB.merge(dfnumTidy[['訂單編號','客戶信箱']],on = '訂單編號',how = 'outer')
dfB = dfB.reindex(columns = KindFoodBookingBill.columns)

dfB['客戶手機'] = dfB['客戶手機'].apply(str)
dfB.loc[:,'客戶手機'] = '0' + dfB.loc[:,'客戶手機']

dfB.to_excel("output/KindFoodBooking雲端發票開立.xlsx",index=None)
tqdm.pandas()
dfB.progress_apply(lambda x: x)
reset_col("output/KindFoodBooking雲端發票開立.xlsx")


# In[7]:


import pandas as pd
import openpyxl
from openpyxl import load_workbook



workbook = load_workbook("output/KindFoodBooking雲端發票開立.xlsx")
sheet = workbook.active

for i in range(2,sheet.max_row+1):
  sheet["A"+ format(i)] = "3 "
  sheet["B"+ format(i)] = "0 "
  sheet["F"+ format(i)] = "1 "
  sheet["O"+ format(i)] = "1 "
  sheet["S"+ format(i)] = "1 "
  sheet["T"+ format(i)] = "生鮮食材組 "
  sheet["U"+ format(i)] = "1 "
  sheet["Y"+ format(i)] = "2 "
  sheet["Z"+ format(i)] = "1 "
  sheet["AA"+ format(i)] = "0 "

workbook.save(filename = "output/KindFoodBooking雲端發票開立.xlsx")


# In[8]:


from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBookingTidy = pd.read_excel("output/KindFoodBookingTidy.xlsx")
KindFoodBookingStock = pd.read_excel("input/庫存項目.xlsx",sheet_name="庫存表")
KindFoodBookingCompare = pd.read_excel("output/KindFoodBookingTidy.xlsx",sheet_name="樞紐分析")

dfS = KindFoodBookingStock
dfC = KindFoodBookingCompare

dfC = dfC.rename(columns={'Unnamed: 0' : '商品編號'})
dfS = dfS.rename(columns={'庫存料號':'商品編號'})

dfC = dfC.merge(dfS[['商品編號','庫存數量(僅參考)']],on ='商品編號',how = 'outer')

sumindex = dfC[dfC['商品編號']=='總和'].index.values.astype(int)[0]

for i in range(sumindex+1,len(dfC.index)):
  dfC = dfC.drop(index=i)

with pd.ExcelWriter("output/KindFoodBookingTidy.xlsx") as writer:
  KindFoodBookingTidy.to_excel(writer,sheet_name="訂單明細",index=None)
  dfC.to_excel(writer,sheet_name="樞紐分析",index=None)

tqdm.pandas()
df.progress_apply(lambda x: x)
reset_col("output/KindFoodBookingTidy.xlsx")


# In[ ]:




