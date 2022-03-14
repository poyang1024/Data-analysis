#!/usr/bin/env python
# coding: utf-8

# In[14]:


print('使用前請先詳閱Readme文件中的使用步驟......')
print()
print('請耐心等候，程式運行完後會自動關閉，於output資料夾匯出所需檔案......')


# In[15]:


from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 
def reset_col(filename):
	wb=load_workbook(filename)
	for sheet in wb.sheetnames:
		ws=wb[sheet]
		df=pd.read_excel(filename,sheet).fillna('-')
		df.loc[len(df)]=list(df.columns)						#把標题行附加到最后一行
		for col in df.columns:				
			index=list(df.columns).index(col)					#列序號
			letter=get_column_letter(index+1)					#列字母
			collen=df[col].apply(lambda x:len(str(x).encode())).max()	#獲取這一列長度的最大值
			ws.column_dimensions[letter].width=collen*0.45+10

	wb.save(filename)


# In[16]:


from tqdm import tqdm
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

KindFoodBookingNewStock = pd.read_excel("input/庫存項目.xlsx",sheet_name="庫存表")
KindFoodBookingOldStock = pd.read_excel("input/安全庫存.xlsx")

dfnew = KindFoodBookingNewStock
dfold = KindFoodBookingOldStock


dfnew = dfnew.rename(columns={'庫存料號':'商品編號'})


dfold = dfold.merge(dfnew[['商品編號','庫存數量(僅參考)']],on ='商品編號',how = 'outer')
for i in range(40,len(dfold.index)):
  dfold = dfold.drop(index=i)

dfold = dfold.rename(columns={'愛上新鮮庫存':'刪掉'})
dfold = dfold.rename(columns={'庫存數量(僅參考)':'愛上新鮮庫存'})

dfold = dfold.drop(columns={'刪掉'})
dfold = dfold.reindex(columns = KindFoodBookingOldStock.columns)
dfold.to_excel("output/KindFoodBooking安全庫存.xlsx",index=None)

tqdm.pandas()
dfold.progress_apply(lambda x: x)
reset_col("output/KindFoodBooking安全庫存.xlsx")


# In[19]:


import pandas as pd
import openpyxl
from openpyxl import load_workbook



workbook = load_workbook("output/KindFoodBooking安全庫存.xlsx")
sheet = workbook.active

sheet["C2"] = "= M2 "
for i in range(2,11):
  sheet["C"+ format(i)] = "= ROUND(M"+ format(i) + ",0)"
for i in range(18,24):
  sheet["C"+ format(i)] = "= ROUND(M"+ format(i) + ",0)"
for i in range(12,18):
  sheet["C"+ format(i)] = "= ROUND((M"+ format(i) + ")" + "/30*14" + ",0)"
for i in range(24,36):
  sheet["C"+ format(i)] = "= ROUND((M"+ format(i) + ")" + "/30*14" + ",0)"
sheet["C36"] = "= ROUND(M36" + ",0)"
sheet["C37"] = "=ROUND((M37/30)*14" + ",0)"

for i in range(38,41):
  sheet["C"+ format(i)] = "= ROUND(M"+ format(i) + ",0)"

for i in range(2,41):
  sheet["D"+ format(i)] = "= ROUND((G"+ format(i) + ")" + "/(M" + format(i)+"/30),0)"

for i in range(2,41):
  sheet["E"+ format(i)] = "= ROUND((G"+ format(i) + "+" +  "H"+ format(i) + "+" + "I"+ format(i) + ")" + "/(M" + format(i)+"/30),0)"

for i in range(2,41):
  sheet["F"+ format(i)] = "= ROUND(IF(I"+ format(i) + ">0,0,IF(C" + format(i) + ">G" + format(i) + "+H" + format(i) + ",C" + format(i) + "*1.3,0)),0)"

sheet.freeze_panes = 'C1'
sheet.freeze_panes = 'C2'

workbook.save(filename = "output/KindFoodBooking安全庫存.xlsx")


# In[ ]:




